from typing import List
from openpyxl import load_workbook
from openpyxl.cell import MergedCell


def get_sheet_object(data_path, sheet='Sheet1', if_data_only=True):
    """       
        Pack the function of obtaining sheet objects
    """
    wb = load_workbook(data_path, data_only=if_data_only)
    Sheet = wb.get_sheet_by_name(sheet) 

    return Sheet


def save_modify_sheet(final_sheet_path,data_path,required_data:List,sheet='Sheet1'):
    """
        Modify the final sheet by target value and save it
    """
    wb = load_workbook(final_sheet_path, data_only=True)
    Final_sheet = wb.get_sheet_by_name(sheet) 
    for data_tuple in required_data:
        for col in range(1,len(data_tuple)):
            Final_sheet.cell(row=data_tuple[0],column=col).value = data_tuple[col]
    
    wb.save(data_path)


def parser_merged_cell(sheet, row, col):
    """
        检查是否为合并单元格并获取对应行列单元格的值。
        如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
        :param sheet: 当前工作表对象
        :param row: 需要获取的单元格所在行
        :param col: 需要获取的单元格所在列
        :return: 
    """
    cell = sheet.cell(row=row, column=col) # 定位到某一单元格
    if isinstance(cell, MergedCell): # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cell_ranges:
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    cell = cell.value.replace(' ','')

    return cell


def get_target_list(data_path_list):
    """
        Get Date;
        Total_Load_Size;
        Mid_To_Long_Term_Total_Load_Size;
        Corporate_Loan_Scale;
        Mid_To_Long_Term_Corporate_Loan_Scale;
        Mid_Loan_Scale;
        Mid_Corporate_Loan_Scale
    """
    Date = []
    Total_Load_Size = []
    Mid_To_Long_Term_Total_Load_Size = []
    Corporate_Loan_Size = []
    Mid_To_Long_Term_Corporate_Loan_Size = []
    for data_path in data_path_list:
        current_sheet = get_sheet_object(data_path)
        date = parser_merged_cell(current_sheet,2,1)
        Total_Load_Size_target = current_sheet['D30'].value
        Mid_To_Long_Term_Total_Load_Size_target = current_sheet['M30'].value
        Corporate_Loan_Scale_target = current_sheet['B30'].value
        Mid_To_Long_Term_Corporate_Loan_Scale_target = current_sheet['H30'].value
        # Collection
        Date.append(date)
        Total_Load_Size.append(Total_Load_Size_target)
        Mid_To_Long_Term_Total_Load_Size.append(Mid_To_Long_Term_Total_Load_Size_target)
        Corporate_Loan_Size.append(Corporate_Loan_Scale_target)
        Mid_To_Long_Term_Corporate_Loan_Size.append(Mid_To_Long_Term_Corporate_Loan_Scale_target)
        Mid_Loan_Scale = [(mid_size/total_size) for mid_size,total_size 
                           in zip(Mid_To_Long_Term_Total_Load_Size,Total_Load_Size)]
        Mid_Corporate_Loan_Scale = [(mid_com_size/total_com_size) for mid_com_size,total_com_size
                                    in zip(Mid_To_Long_Term_Corporate_Loan_Size,Corporate_Loan_Size)]

    return Date,Total_Load_Size, Mid_To_Long_Term_Total_Load_Size, \
           Corporate_Loan_Size, Mid_To_Long_Term_Corporate_Loan_Size,\
           Mid_Loan_Scale, Mid_Corporate_Loan_Scale